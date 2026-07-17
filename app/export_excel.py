"""
Формирование Excel-отчёта.
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path

from paths import REPORTS_DIR

RISK_FILL = {
    "high":          PatternFill("solid", fgColor="FFD32F2F"),
    "medium":        PatternFill("solid", fgColor="FFF57C00"),
    "low":           PatternFill("solid", fgColor="FF388E3C"),
    "informational": PatternFill("solid", fgColor="FF1976D2"),
}
RISK_LABELS = {
    "high": "Высокий",
    "medium": "Средний",
    "low": "Низкий",
    "informational": "Информационный",
}
LEGAL_LABELS = {
    "not_reviewed": "Не проверено",
    "risk_confirmed": "Риск подтверждён",
    "risk_not_confirmed": "Риск не подтверждён",
    "archived": "Архив",
}
SOURCE_LABELS = {
    "kz_registry": "Реестр KZ",
    "kz_bulletin": "Бюллетень KZ",
    "wipo": "WIPO",
    "madrid": "Madrid",
}
OBJECT_LABELS = {
    "trademark": "Товарный знак",
    "well_known": "Общеизвестный ТЗ",
}
STATUS_LABELS = {
    "active": "Действует",
    "application": "Заявка",
    "expired": "Прекращён",
    "refused": "Отказ",
    "unknown": "Иной",
}


def generate_report(
    marks: list[dict],
    title: str,
    period_from: str,
    period_to: str,
    profiles: list[str],
    sources: list[str],
) -> str:
    """
    Создаёт Excel-файл и возвращает путь к нему.
    """
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = REPORTS_DIR / filename

    wb = openpyxl.Workbook()

    _create_summary_sheet(wb.active, marks, title, period_from, period_to, profiles, sources)
    wb.active.title = "Сводка"

    _create_marks_sheet(wb, marks, "Все записи")
    _create_marks_sheet(wb, [m for m in marks if m.get("risk_level") == "high"], "Высокий риск")
    _create_marks_sheet(wb, [m for m in marks if m.get("risk_level") == "medium"], "Средний риск")
    _create_marks_sheet(
        wb,
        [m for m in marks if m.get("legal_status") == "not_reviewed"],
        "Требуют проверки",
    )

    wb.save(str(filepath))
    return str(filepath)


def _header_style():
    return {
        "font": Font(bold=True, color="FFFFFFFF", size=10),
        "fill": PatternFill("solid", fgColor="FF37474F"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
    }


def _apply_header(cell, value):
    cell.value = value
    h = _header_style()
    cell.font = h["font"]
    cell.fill = h["fill"]
    cell.alignment = h["alignment"]


def _thin_border():
    thin = Side(style="thin", color="FFB0BEC5")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _create_summary_sheet(ws, marks, title, period_from, period_to, profiles, sources):
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    ws.merge_cells("A1:B1")
    cell = ws["A1"]
    cell.value = title
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal="center")

    rows = [
        ("Период мониторинга", f"{period_from} — {period_to}"),
        ("Источники", ", ".join(SOURCE_LABELS.get(s, s) for s in sources)),
        ("Профили", ", ".join(profiles)),
        ("Дата формирования", datetime.now().strftime("%d.%m.%Y %H:%M")),
        ("", ""),
        ("Показатель", "Значение"),
        ("Всего найдено записей", len(marks)),
        ("Высокий риск", sum(1 for m in marks if m.get("risk_level") == "high")),
        ("Средний риск", sum(1 for m in marks if m.get("risk_level") == "medium")),
        ("Низкий риск", sum(1 for m in marks if m.get("risk_level") == "low")),
        ("Информационные", sum(1 for m in marks if m.get("risk_level") == "informational")),
        ("Не проверено юристом", sum(1 for m in marks if m.get("legal_status") == "not_reviewed")),
        ("Риск подтверждён", sum(1 for m in marks if m.get("legal_status") == "risk_confirmed")),
        ("Включено в отчёт", sum(1 for m in marks if m.get("include_in_report"))),
    ]

    for i, (label, value) in enumerate(rows, start=2):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = value
        if label == "Показатель":
            _apply_header(ws[f"A{i}"], label)
            _apply_header(ws[f"B{i}"], value)


def _create_marks_sheet(wb, marks, sheet_name):
    ws = wb.create_sheet(title=sheet_name[:31])

    headers = [
        "Риск", "Обозначение", "Тип объекта", "Источник",
        "№ заявки", "№ регистрации", "Дата заявки", "Дата регистрации",
        "Дата публикации", "Классы МКТУ", "Правообладатель",
        "Статус знака", "Причина совпадения", "Юридический статус",
        "Комментарий юриста", "Рекомендация", "В отчёт", "Ссылка",
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        _apply_header(cell, header)

    col_widths = [12, 22, 18, 14, 14, 16, 13, 16, 16, 14, 25, 13, 35, 18, 30, 18, 8, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 30

    for row_idx, mark in enumerate(marks, 2):
        risk = mark.get("risk_level", "informational")
        classes_raw = mark.get("nice_classes_str", "") or mark.get("nice_classes", "")

        row_data = [
            RISK_LABELS.get(risk, risk),
            mark.get("designation", ""),
            OBJECT_LABELS.get(mark.get("object_type"), mark.get("object_type", "")),
            SOURCE_LABELS.get(mark.get("source_code"), mark.get("source_code", "")),
            mark.get("application_number", ""),
            mark.get("registration_number", ""),
            mark.get("application_date", ""),
            mark.get("registration_date", ""),
            mark.get("publication_date", ""),
            str(classes_raw),
            mark.get("owner", ""),
            STATUS_LABELS.get(mark.get("status_mark"), mark.get("status_mark", "")),
            mark.get("match_reason", ""),
            LEGAL_LABELS.get(mark.get("legal_status"), mark.get("legal_status", "")),
            mark.get("lawyer_comment", ""),
            mark.get("recommended_action", ""),
            "Да" if mark.get("include_in_report") else "Нет",
            mark.get("source_url", ""),
        ]

        fill = RISK_FILL.get(risk)
        border = _thin_border()

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=(col_idx in (13, 15)))
            if col_idx == 1 and fill:
                cell.fill = fill
                cell.font = Font(color="FFFFFFFF", bold=True, size=9)

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions
